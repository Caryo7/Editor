from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Protection
from openpyxl.styles import Alignment
from lg import *
import os

PATH_PROG = '.'

def export():
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)

    lgs = {
        'es': {'title': 'Espagnol', 'wb': None, 'dic': {}},
        'it': {'title': 'Italien', 'wb': None, 'dic': {}},
        'ch': {'title': 'Chinois', 'wb': None, 'dic': {}},
        'al': {'title': 'Allemand', 'wb': None, 'dic': {}},
        'an': {'title': 'Anglais', 'wb': None, 'dic': {}},
        'fr': {'title': 'Français', 'wb': None, 'dic': {}},
        }

    l = Lg('fr', 'utf-8')
    fr = list(l.get().values()).copy()

    for lg in list(lgs.keys()):
        lgs[lg]['wb'] = wb.create_sheet(lgs[lg]['title'])
        lgs[lg]['wb'].title = lgs[lg]['title']

        lgs[lg]['wb'].cell(row = 1, column = 1, value = 'Appelant')
        lgs[lg]['wb'].cell(row = 1, column = 2, value = 'Français')
        lgs[lg]['wb'].cell(row = 1, column = 3, value = lgs[lg]['title'])

        lgs[lg]['wb'].column_dimensions['A'].width = 30
        lgs[lg]['wb'].column_dimensions['A'].hidden= True
        lgs[lg]['wb'].column_dimensions['B'].width = 50
        lgs[lg]['wb'].column_dimensions['C'].width = 50

        l = Lg(lg, 'utf-8')
        lgs[lg]['dic'] = l.get()

        row = 1
        n = 0
        for caller, word in lgs[lg]['dic'].items():
            row += 1
            if lg == 'fr':
                lgs[lg]['wb'].cell(row = row, column = 3, value = fr[n])
            lgs[lg]['wb'].cell(row = row, column = 2, value = fr[n])
            lgs[lg]['wb'].cell(row = row, column = 1, value = caller)
            n += 1

    
        lgs[lg]['wb'].protection.sheet = True
        for col in ['C']:
            for cell in lgs[lg]['wb'][col]:
                if cell.row > 1:
                    cell.protection = Protection(locked=False)

        for col in ['A', 'B', 'C']:
            for cell in lgs[lg]['wb'][col]:
                cell.alignment = Alignment(wrapText = True)

        if lg == 'fr':
            lgs[lg]['wb'].sheet_state = 'hidden'

    wb.save(PATH_PROG + "/temp/export_langs.xlsx")
    os.popen(PATH_PROG + '/temp/export_langs.xlsx')

def _import():
    wb = load_workbook(PATH_PROG + '/temp/export_langs.xlsx')
    names = {
        'Espagnol': 'es',
        'Italien': 'it',
        'Chinois': 'ch',
        'Allemand': 'al',
        'Anglais': 'an',
        'Français': 'fr',
        }

    for sn in wb.sheetnames:
        ws = wb[sn]
        dic = {}
        for row in ws:
            if row[0].value != 'Appelant':
                dic[row[0].value] = row[2].value

        f = open(PATH_PROG + '/temp/' + names[sn] + '.lang', 'w')
        f.write('[' + names[sn] + ']\n')
        for k, v in dic.items():
            if v == None:
                v = ''

            f.write(k + ' = ' + v + '\n')
        f.close()

print('1. Exporter sous Excel')
print('2. Importer depuis Excel')
choix = int(input('> '))
if choix == 1:
    export()
else:
    _import()
